#!/usr/bin/env python3
"""
PDF ve Excel çıktıları üretir:
  - assets/kupa-sunum.pdf — landscape A4, her sayfada Mavi-Crossover-Sports-Logo
    (üst-sağ köşede), sadece kapak sayfasında eski header-logo.png ortada.
    Kupa Duo bölümleri ve entegrasyon thumbları YouTube linki ile tıklanabilir.
  - assets/kupa-paketler.xlsx — Diğer Bütçeler + Paketler sheets.
    Sol: Kırmızı-Transparan logo, sağ: Mavi-Crossover-Sports-Logo.
    Diğer Bütçeler'de sütun "Bütçe" + "+KDV", Paketler'de eski fiyat sütunu yok.
"""

from pathlib import Path

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm, inch
from reportlab.lib.colors import HexColor, black, white
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle,
    PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
ASSETS = ROOT / "assets"
IMG = ASSETS / "img"
FONTS = ASSETS / "fonts"
HERO_LOGO_OLD = ROOT / "header-logo.png"               # 6 kareli, sadece PDF kapak sayfası ortasında
LOGO_MAVI = ROOT / "Mavi-Crossover-Sports-Logo.png"   # 1080x1080 kare
LOGO_KIRMIZI = ROOT / "Kırmızı-Transparan.png"         # 1200x960
D20_LOGO = IMG / "hero-d20-4310a162.png"               # PDF kapak footer
DUO_THUMB = ROOT / "kupaduo1-son-sunum.jpeg"
MASASI_THUMB = ROOT / "kupamasasi1-son-sunum.jpeg"

# ─── D20 marka fontu (AllRoundGothic) — assets/fonts/ ───────────────────────
def _register_fonts():
    files = {
        'TR':            FONTS / 'AllRoundGothic-Book.ttf',
        'TR-Bold':       FONTS / 'AllRoundGothic-Bold.ttf',
        'TR-Italic':     FONTS / 'AllRoundGothic-BookOblique.ttf',
        'TR-BoldItalic': FONTS / 'AllRoundGothic-BoldOblique.ttf',
    }
    if not files['TR'].exists():
        return False
    for name, path in files.items():
        if path.exists():
            pdfmetrics.registerFont(TTFont(name, str(path)))
    from reportlab.pdfbase.pdfmetrics import registerFontFamily
    registerFontFamily('TR',
                       normal='TR',
                       bold='TR-Bold',
                       italic='TR-Italic',
                       boldItalic='TR-BoldItalic')
    return True


_HAS_TR_FONT = _register_fonts()
FONT_REG = 'TR' if _HAS_TR_FONT else 'Helvetica'
FONT_BOLD = 'TR-Bold' if _HAS_TR_FONT else 'Helvetica-Bold'

# ─── Brand colors ────────────────────────────────────────────────────────────
GOLD = HexColor("#d4a843")
GOLD_LIGHT = HexColor("#f0d078")
RED = HexColor("#e74c3c")
BLUE = HexColor("#2980b9")
BG = HexColor("#0f1115")
CARD = HexColor("#1a1d24")
BORDER = HexColor("#2a2d35")
TEXT = HexColor("#e8e6e1")
TEXT_MUTED = HexColor("#8a8d94")
LINK = HexColor("#f0d078")  # gold-light for links

# ─── PDF — PowerPoint classic 4:3 (10" × 7.5") ───────────────────────────────
PPT_CLASSIC = (10 * inch, 7.5 * inch)
PAGE_W, PAGE_H = PPT_CLASSIC          # 25.4 × 19.05 cm
LMARGIN = 14 * mm
RMARGIN = 14 * mm
TMARGIN = 22 * mm                      # gold accent line + boşluk için
BMARGIN = 16 * mm


def _header_footer(canvas, doc):
    """Sayfa 1: D20 logo footer'da ortada. Diğer sayfalar: sadece çizgi+metin."""
    canvas.saveState()

    # Background
    canvas.setFillColor(BG)
    canvas.rect(0, 0, PAGE_W, PAGE_H, stroke=0, fill=1)

    # Top & bottom accent lines
    canvas.setStrokeColor(GOLD)
    canvas.setLineWidth(1.0)
    canvas.line(LMARGIN, PAGE_H - 8 * mm, PAGE_W - RMARGIN, PAGE_H - 8 * mm)
    canvas.setStrokeColor(BORDER)
    canvas.setLineWidth(0.4)
    canvas.line(LMARGIN, BMARGIN - 4, PAGE_W - RMARGIN, BMARGIN - 4)

    # Top label (sadece sayfa 2+)
    canvas.setFillColor(GOLD_LIGHT)
    canvas.setFont(FONT_BOLD, 9)
    if doc.page > 1:
        canvas.drawString(LMARGIN, PAGE_H - 6 * mm, "KUPA DUO & KUPA MASASI · İçerik Planı")

    # Page 1 footer: D20 logo center (önceki boyutun %70'i)
    if doc.page == 1 and D20_LOGO.exists():
        try:
            from PIL import Image as PILImage
            with PILImage.open(D20_LOGO) as im:
                w, h = im.size
            ratio = h / w
        except Exception:
            ratio = 0.72
        d20_w = 15.4 * mm                      # 22mm × 0.7
        d20_h = d20_w * ratio
        try:
            canvas.drawImage(
                str(D20_LOGO),
                (PAGE_W - d20_w) / 2, BMARGIN / 2 - d20_h / 2 + 1 * mm,
                width=d20_w, height=d20_h, mask='auto',
            )
        except Exception:
            pass

    # Footer text "D20-Crossover Sports" + sayfa no
    canvas.setFillColor(TEXT_MUTED)
    canvas.setFont(FONT_REG, 9)
    canvas.drawString(LMARGIN, BMARGIN - 11, "D20-Crossover Sports")
    canvas.drawRightString(PAGE_W - RMARGIN, BMARGIN - 11, f"Sayfa {doc.page}")

    canvas.restoreState()


def _styles():
    base = getSampleStyleSheet()
    return {
        'title':   ParagraphStyle('title', parent=base['Heading1'],
                                  fontName=FONT_BOLD, fontSize=24,
                                  textColor=GOLD_LIGHT, leading=28, spaceAfter=3 * mm),
        'h2':      ParagraphStyle('h2', parent=base['Heading2'],
                                  fontName=FONT_BOLD, fontSize=16,
                                  textColor=GOLD, leading=20, spaceBefore=4 * mm,
                                  spaceAfter=2 * mm),
        'h3':      ParagraphStyle('h3', parent=base['Heading3'],
                                  fontName=FONT_BOLD, fontSize=13,
                                  textColor=TEXT, leading=16, spaceAfter=1 * mm),
        'body':    ParagraphStyle('body', parent=base['Normal'],
                                  fontName=FONT_REG, fontSize=12,
                                  textColor=TEXT, leading=15, spaceAfter=2 * mm),
        'body_muted': ParagraphStyle('body_muted', parent=base['Normal'],
                                  fontName=FONT_REG, fontSize=11,
                                  textColor=TEXT_MUTED, leading=14, spaceAfter=2 * mm),
        'small':   ParagraphStyle('small', parent=base['Normal'],
                                  fontName=FONT_REG, fontSize=10,
                                  textColor=TEXT, leading=13),
        'small_muted': ParagraphStyle('small_muted', parent=base['Normal'],
                                  fontName=FONT_REG, fontSize=10,
                                  textColor=TEXT_MUTED, leading=13),
        'link':    ParagraphStyle('link', parent=base['Normal'],
                                  fontName=FONT_REG, fontSize=10,
                                  textColor=LINK, leading=13),
    }


def _table(data, col_widths=None):
    t = Table(data, colWidths=col_widths)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GOLD),
        ('TEXTCOLOR', (0, 0), (-1, 0), black),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 1), (-1, -1), CARD),
        ('TEXTCOLOR', (0, 1), (-1, -1), TEXT),
        ('FONTNAME', (0, 1), (-1, -1), FONT_REG),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 0.4, BORDER),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ])
    t.setStyle(style)
    return t


def _link(url: str, label: str) -> str:
    """Reportlab Paragraph içine clickable link inline'ı döner."""
    return f'<link href="{url}" color="#f0d078"><u>{label}</u></link>'


def build_pdf(out_path: Path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=PPT_CLASSIC,
        leftMargin=LMARGIN, rightMargin=RMARGIN,
        topMargin=TMARGIN, bottomMargin=BMARGIN,
        title="Kupa Duo & Kupa Masası · İçerik Planı",
        author="D20 Media House",
    )
    s = _styles()
    story = []

    # ── PAGE 1: Cover — sıkıştırılmış, tek sayfaya sığar ─────────────────────
    story.append(Spacer(1, 4 * mm))
    if HERO_LOGO_OLD.exists():
        try:
            from PIL import Image as PILImage
            with PILImage.open(HERO_LOGO_OLD) as im:
                w, h = im.size
            target_w = 95 * mm                         # 130 → 95mm (kompakt)
            img = Image(str(HERO_LOGO_OLD), width=target_w, height=target_w * h / w)
            img.hAlign = 'CENTER'
            story.append(img)
        except Exception:
            pass
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("HEDEF DÜNYA KUPASI!", ParagraphStyle(
        'cover_title', fontName=FONT_BOLD, fontSize=26,
        textColor=RED, alignment=TA_CENTER, leading=30)))
    story.append(Paragraph("Futbol ve Eğlence", ParagraphStyle(
        'cover_sub', fontName=FONT_REG, fontSize=16,
        textColor=TEXT, alignment=TA_CENTER, leading=20, spaceBefore=1 * mm)))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(
        "2026 Dünya Kupası'nda analiz her yerde. "
        "Futbol ve eğlenceyi bir arada yaşatan tek masa burada!",
        ParagraphStyle('cover_lead', fontName=FONT_BOLD, fontSize=11,
                       textColor=TEXT, alignment=TA_CENTER, leading=14)))
    story.append(Paragraph(
        f"Crossover Sports — {_link('https://www.youtube.com/@Crossover-Sports', 'youtube.com/@Crossover-Sports')}",
        ParagraphStyle('cover_desc', fontName=FONT_REG, fontSize=10,
                       textColor=TEXT_MUTED, alignment=TA_CENTER, leading=13, spaceBefore=2 * mm)))
    story.append(PageBreak())

    # ── PAGE 2: Show Cards (yatay 2 kolon) ──────────────────────────────────
    story.append(Paragraph("ŞOVLAR", s['title']))
    story.append(Paragraph("İki faz · iki format · tek marka", s['body_muted']))
    story.append(Spacer(1, 3 * mm))

    # Side-by-side
    img_w = 100 * mm
    img_h = img_w * 9 / 16
    duo_img = Image(str(DUO_THUMB), width=img_w, height=img_h) if DUO_THUMB.exists() else Paragraph("", s['body'])
    masasi_img = Image(str(MASASI_THUMB), width=img_w, height=img_h) if MASASI_THUMB.exists() else Paragraph("", s['body'])

    duo_text = (
        '<font color="#e74c3c"><b>FAZ 1 — KUPA DUO</b></font>  ·  YAYINDA<br/><br/>'
        'Dünya Kupası boyunca maç üzerinden aşırı eğlenceli muhabbetler '
        '<b>Mustafa Demirtaş</b> ve <b>Koray Koç</b> ile Kupa DUO\'da!'
    )
    masasi_text = (
        '<font color="#2980b9"><b>FAZ 2 — KUPA MASASI</b></font>  ·  06 Haziran – 20 Temmuz<br/><br/>'
        'Haziran\'da futbol ve eğlence dozu 2 katına çıkıyor. '
        '<b>Mustafa Demirtaş</b>, <b>Koray Koç</b>, <b>Buğra Oflaz</b> ve <b>Nazmi Sinan Mıhçı</b> ile.'
    )
    show_cards = Table(
        [
            [duo_img, masasi_img],
            [Paragraph(duo_text, s['small']), Paragraph(masasi_text, s['small'])],
        ],
        colWidths=[img_w + 4 * mm, img_w + 4 * mm],
    )
    show_cards.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(show_cards)
    story.append(PageBreak())

    # ── PAGE 3: Kupa Duo Episodes (clickable links) ─────────────────────────
    story.append(Paragraph("KUPA DUO — YAYINDA & GELECEK BÖLÜMLER", s['title']))
    story.append(Paragraph("Toplam 8 bölüm · Pazar yayınları", s['body_muted']))
    story.append(Spacer(1, 3 * mm))

    duo_episodes = [
        ('01', '—', 'EN İYİ DÜNYA KUPASI ŞARKILARI · ARDA GÜLER vs. KENAN YILDIZ · TARKAN · SHAKIRA · DUA LIPA',
         'YAYINDA', 'https://www.youtube.com/watch?v=JgCTOBnsmeA&t=2s'),
        ('02', '—', "MAÇINI KAÇIRMAYACAĞIMIZ YILDIZLAR · LAMINE YAMAL · MESSI vs. RONALDO · BBG'DEKİ SIZMACI",
         'YAYINDA', 'https://www.youtube.com/watch?v=0zvJTQ_jtr0'),
        ('03', '—', 'EN İYİ DÜNYA KUPASI STİLLERİ · ÜMİT DAVALA SAÇI · RONALDO ÇİZİĞİ · BAGGIO · PSG-BAYERN',
         'YAYINDA', 'https://www.youtube.com/watch?v=JOUingQfLXI&t=115s'),
        ('04', '10 Mayıs Pazar', '—', 'Gelecek Bölüm', None),
        ('05', '17 Mayıs Pazar', '—', 'Gelecek Bölüm', None),
        ('06', '24 Mayıs Pazar', '—', 'Gelecek Bölüm', None),
        ('07', '31 Mayıs Pazar', '—', 'Gelecek Bölüm', None),
        ('08', '02 Haziran Salı', '—', 'Gelecek Bölüm', None),
    ]

    duo_rows = [['#', 'Tarih', 'Bölüm Başlığı', 'Durum']]
    for ep in duo_episodes:
        num, date, title, status, url = ep
        if url:
            title_p = Paragraph(_link(url, title), s['small'])
        else:
            title_p = Paragraph(title, s['small_muted'])
        duo_rows.append([
            Paragraph(num, s['small']),
            Paragraph(date, s['small']),
            title_p,
            Paragraph(status, s['small']),
        ])
    story.append(_table(duo_rows, col_widths=[10 * mm, 26 * mm, 152 * mm, 32 * mm]))
    story.append(PageBreak())

    # ── PAGE 4: Kupa Masası Schedule ────────────────────────────────────────
    story.append(Paragraph("KUPA MASASI — YAYIN TAKVİMİ", s['title']))
    story.append(Paragraph("Toplam 13 bölüm · 06 Haziran – 20 Temmuz", s['body_muted']))
    story.append(Spacer(1, 3 * mm))
    masasi_ep = [
        ['#', 'Tarih', 'Açıklama'],
        ['01', '06 Haziran (Cts)', 'Preview 1'],
        ['02', '10 Haziran (Çrş)', 'Preview 2'],
        ['03', '14 Haziran (Paz)', 'TUR-AUS maç günü'],
        ['04', '17 Haziran (Çrş)', '—'],
        ['05', '20 Haziran (Cts)', 'TUR-PAR maç günü'],
        ['06', '23 Haziran (Salı)', '—'],
        ['07', '26 Haziran (Cuma)', 'TUR-USA maç günü'],
        ['08', '30 Haziran (Salı)*', '—'],
        ['09', '04 Temmuz (Cts)*', '—'],
        ['10', '08 Temmuz (Çrş)*', '—'],
        ['11', '12 Temmuz (Paz)*', '—'],
        ['12', '16 Temmuz (Prş)*', '—'],
        ['13', '20 Temmuz (Pts)', 'Final/Kapanış'],
    ]
    masasi_p = [[Paragraph(c, s['small']) for c in row] for row in masasi_ep]
    story.append(_table(masasi_p, col_widths=[12 * mm, 52 * mm, 156 * mm]))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph("* Tarih takım performansına göre değişebilir", s['small_muted']))
    story.append(PageBreak())

    # ── PAGE 5: Özet Toplamlar + Bütçe ──────────────────────────────────────
    story.append(Paragraph("ÖZET TOPLAMLAR — KUPA MASASI", s['title']))
    story.append(Paragraph("13 bölüm üzerinden tahmini içerik & görüntüleme", s['body_muted']))
    story.append(Spacer(1, 3 * mm))
    summary = [
        ['Format', 'Adet', 'Tahmini Görüntüleme'],
        ['YouTube', '13', '~2.93M'],
        ['Shorts',  '39', '~780K'],
        ['Reels',   '39', '~1.56M'],
        ['TikTok',  '39', '~390K'],
        ['TOPLAM',  '130 İÇERİK', '~5.66M'],
    ]
    sum_p = [[Paragraph(c, s['small']) for c in row] for row in summary]
    story.append(_table(sum_p, col_widths=[70 * mm, 70 * mm, 80 * mm]))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("BÜTÇE & CPV", s['h2']))
    cpv = [
        ['Kategori', 'Değer'],
        ['Toplam Bütçe (Entegrasyon)', '2.600.000 +KDV'],
        ['Genel CPV (Cost Per View)', '0,46'],
    ]
    cpv_p = [[Paragraph(c, s['small']) for c in row] for row in cpv]
    story.append(_table(cpv_p, col_widths=[130 * mm, 90 * mm]))
    story.append(PageBreak())

    # ── PAGE 6: Kanal İstatistikleri + Demografi ────────────────────────────
    story.append(Paragraph("CROSSOVER SPORTS — KANAL İSTATİSTİKLERİ", s['title']))
    story.append(Paragraph("Ömür boyu: 22 May 2024 – 6 May 2026", s['body_muted']))
    story.append(Spacer(1, 3 * mm))
    stats = [
        ['Görüntülenme', 'İzlenme Süresi', 'Aboneler'],
        ['5,6M',          '514,2B Saat',    '43,8B'],
    ]
    stats_p = [[Paragraph(c, s['small']) for c in row] for row in stats]
    story.append(_table(stats_p, col_widths=[73 * mm, 73 * mm, 73 * mm]))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("DEMOGRAFİ", s['h2']))
    demog = [
        ['Yaş Grubu', 'Oran',     'Cinsiyet', 'Oran'],
        ['25-34',     '%40,5',    'Erkek',    '%93,6'],
        ['35-44',     '%25,0',    'Kadın',    '%6,4'],
        ['18-24',     '%18,8',    '',         ''],
        ['45-54',     '%9,8',     '',         ''],
        ['55-64',     '%4,2',     '',         ''],
        ['65+',       '%1,7',     '',         ''],
    ]
    demog_p = [[Paragraph(c, s['small']) for c in row] for row in demog]
    story.append(_table(demog_p, col_widths=[52 * mm, 52 * mm, 52 * mm, 64 * mm]))
    story.append(PageBreak())

    # ── PAGE 7: Entegrasyon Çeşitleri (clickable thumbs) ────────────────────
    story.append(Paragraph("ENTEGRASYON ÇEŞİTLERİ", s['title']))
    story.append(Paragraph("Crossover Sports kanalında daha önce uygulanmış formatlar (örnekler tıklanabilir)",
                            s['body_muted']))
    story.append(Spacer(1, 3 * mm))
    integ = [
        ['Format', 'Marka Örnekleri'],
        ['ALT BANT', ' · '.join([
            _link('https://youtu.be/d-IKf3lxXAI?si=72_RYvQWYXfBkpDt&t=417', 'Toyzz Shop'),
            _link('https://www.youtube.com/watch?v=OonL4-zkbAw&t=492s', 'Getiraraç'),
            _link('https://www.youtube.com/watch?v=u-BuHb47h3c', 'S Sport Plus'),
            _link('https://youtu.be/EOdku2cexUM?feature=shared&t=506', 'Cambly'),
        ])],
        ['ÜRÜN YERLEŞTİRME', ' · '.join([
            _link('https://youtu.be/d-IKf3lxXAI?si=aDEKE5HfpV57Xa8-&t=442', 'Topps'),
            _link('https://www.youtube.com/watch?v=Bd9mcvleMMs&t=1604s', 'OMO'),
            _link('https://youtu.be/SDCZoamCRUs?feature=shared&t=15', 'Doritos'),
            _link('https://youtu.be/SDCZoamCRUs?feature=shared&t=88', 'Doritos'),
            _link('https://youtu.be/udMiI2Kxwt8?feature=shared&t=1845', 'HP'),
            _link('https://youtu.be/j7OazTl4D4c?feature=shared', '+1'),
        ])],
        ['AKSİYON ENTEGRE', ' · '.join([
            _link('https://youtu.be/OonL4-zkbAw?feature=shared&t=88', 'Getiraraç'),
            _link('https://youtu.be/OonL4-zkbAw?feature=shared&t=46', 'Getiraraç'),
            _link('https://youtu.be/d-IKf3lxXAI?si=xnMgu5nDWsy0fcks&t=8', 'Toyzz Shop'),
            _link('https://youtu.be/d-IKf3lxXAI?si=09gcZ-hGjbXhLUhR&t=91', 'Topps'),
        ])],
        ['SEPARATÖR', ' · '.join([
            _link('https://www.youtube.com/watch?v=Bd9mcvleMMs&t=1604s', 'OMO'),
            _link('https://youtu.be/udMiI2Kxwt8?feature=shared&t=1755', 'HP'),
        ])],
        ['İNDİRİM KODU', _link('https://www.youtube.com/watch?v=u-BuHb47h3c', 'S Sport Plus')],
    ]
    integ_rows = [[Paragraph(integ[0][0], s['small']), Paragraph(integ[0][1], s['small'])]]
    for row in integ[1:]:
        integ_rows.append([Paragraph(row[0], s['small']), Paragraph(row[1], s['small'])])
    story.append(_table(integ_rows, col_widths=[52 * mm, 168 * mm]))
    story.append(PageBreak())

    # ── PAGE 8: Diğer Bütçeler ──────────────────────────────────────────────
    story.append(Paragraph("DİĞER BÜTÇELER", s['title']))
    story.append(Paragraph("4 format · münferit fiyatlar (13 bölüm)", s['body_muted']))
    story.append(Spacer(1, 3 * mm))
    budgets = [
        ['Format', 'Bütçe', 'Açıklama'],
        ['ENTEGRASYON', '2.600.000 +KDV',
         "Marka bölümün içine giriyor. Konuşuluyor, yaşanıyor, parça oluyor. Reklam değil — sohbet."],
        ['SUNDU / SUNAR', '1.820.000 +KDV',
         '"X sunar" diyoruz, bölüm seninle başlıyor. Rakip giremez. 13 bölüm = 13 imza.'],
        ['ÜRÜN YERLEŞTİRME', '1.040.000 +KDV',
         'Marka kadrajda yaşıyor. Konuşmuyor ama orada. İzleyici farkında olmadan alışıyor.'],
        ['ALTBANT', '650.000 +KDV',
         'Ekran altında akan logo, slogan, çağrı. Sade ama 13 bölüm tekrar = akılda kalır.'],
    ]
    budget_p = [[Paragraph(c, s['small']) for c in row] for row in budgets]
    story.append(_table(budget_p, col_widths=[48 * mm, 44 * mm, 128 * mm]))
    story.append(PageBreak())

    # ── PAGE 9: Paketler ────────────────────────────────────────────────────
    story.append(Paragraph("PAKETLER", s['title']))
    story.append(Paragraph("4 paket · birleşik avantaj fiyatı (13 bölüm)", s['body_muted']))
    story.append(Spacer(1, 3 * mm))
    packages = [
        ['Paket', 'İçerik', 'Bütçe', 'Açıklama'],
        ['BRONZ\n"Görünürlük"',
         'Altbant + Ürün Yerleştirme', '1.250.000 +KDV',
         'Marka 13 bölüm boyunca hem ekran altında hem kadrajda. Sürekli görünürlük.'],
        ['ALTIN\n"Sahiplen"',
         'Sundu + Altbant + Ürün Yerleştirme', '2.320.000 +KDV',
         'Bölümün sahibi sensin. "X sunar" + ürün sahnede + altbant akıyor.'],
        ['ÖZEL\n"Konuş & Sahiplen"',
         'Entegrasyon + Sundu', '3.350.000 +KDV',
         'En güçlü iki silah. Hem bölümün sahibi ol, hem içinde konuşul. Premium.'],
        ['ELMAS\n"Tam Kapsama"',
         'Entegrasyon + Sundu + Altbant + Ürün Yerleştirme',
         '3.970.000 +KDV',
         'Hepsi senin. Bölüm seninle açılıyor, içinde konuşuluyor, kadrajda yaşıyor.'],
    ]
    pkg_p = [[Paragraph(c.replace('\n', '<br/>'), s['small']) for c in row] for row in packages]
    story.append(_table(pkg_p, col_widths=[34 * mm, 65 * mm, 36 * mm, 85 * mm]))
    story.append(PageBreak())

    # ── PAGE 10: Genel Toplam ───────────────────────────────────────────────
    story.append(Paragraph("GENEL TOPLAM — KUPA MASASI", s['title']))
    story.append(Paragraph("Faz 2 — Entegrasyon formatı baz alınmıştır", s['body_muted']))
    story.append(Spacer(1, 3 * mm))
    grand = [
        ['',                'YouTube', 'Shorts', 'Reels', 'TikTok', 'TOPLAM'],
        ['Adet',            '13',      '39',     '39',    '39',     '130 içerik'],
        ['Tahmini İzlenme', '~2.93M',  '~780K',  '~1.56M','~390K',  '~5.66M'],
    ]
    grand_p = [[Paragraph(c, s['small']) for c in row] for row in grand]
    story.append(_table(grand_p, col_widths=[35 * mm, 30 * mm, 30 * mm, 30 * mm, 30 * mm, 65 * mm]))
    story.append(Spacer(1, 3 * mm))
    final = [
        ['', 'Değer'],
        ['Toplam Bütçe (Entegrasyon)', '2.600.000 +KDV'],
        ['Genel CPV', '0,46'],
    ]
    final_p = [[Paragraph(c, s['small']) for c in row] for row in final]
    # Üstteki tablo ile aynı en (220mm): 130 + 90
    story.append(_table(final_p, col_widths=[130 * mm, 90 * mm]))

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)


# ─── EXCEL ───────────────────────────────────────────────────────────────────
def build_excel(out_path: Path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # Sheet 1: Diğer Bütçeler
    ws1 = wb.active
    ws1.title = "Diğer Bütçeler"
    _excel_header(ws1, "DİĞER BÜTÇELER · 4 format · 13 bölüm", n_cols=3)
    headers = ['Format', 'Bütçe', 'Açıklama']
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=5, column=col, value=h)
        c.font = Font(bold=True, color="181818", size=11)
        c.fill = PatternFill("solid", fgColor="d4a843")
        c.alignment = Alignment(horizontal="left", vertical="center")
    rows = [
        ('ENTEGRASYON', '₺2.600.000 +KDV',
         "Marka bölümün içine giriyor. Konuşuluyor, yaşanıyor, parça oluyor. Reklam değil — sohbet."),
        ('SUNDU / SUNAR', '₺1.820.000 +KDV',
         '"X sunar" diyoruz, bölüm seninle başlıyor. Rakip giremez. 13 bölüm = 13 imza.'),
        ('ÜRÜN YERLEŞTİRME', '₺1.040.000 +KDV',
         "Marka kadrajda yaşıyor. Konuşmuyor ama orada. İzleyici farkında olmadan alışıyor."),
        ('ALTBANT', '₺650.000 +KDV',
         "Ekran altında akan logo, slogan, çağrı. Sade ama 13 bölüm tekrar = akılda kalır."),
    ]
    for i, (fmt, price, desc) in enumerate(rows, start=6):
        ws1.cell(row=i, column=1, value=fmt).font = Font(bold=True, size=11)
        cp = ws1.cell(row=i, column=2, value=price)
        cp.font = Font(bold=True, color="d4a843", size=11)
        cp.alignment = Alignment(horizontal="right")
        ws1.cell(row=i, column=3, value=desc).alignment = Alignment(wrap_text=True, vertical="top")
        ws1.row_dimensions[i].height = 38
    _excel_columns(ws1, [25, 22, 70])
    _excel_borders(ws1, 5, 9, 1, 3)

    # Sheet 2: Paketler (eski fiyat sütunu YOK, "Yeni Fiyat" → "Bütçe")
    ws2 = wb.create_sheet("Paketler")
    _excel_header(ws2, "PAKETLER · 4 paket · birleşik avantaj fiyatı · 13 bölüm", n_cols=4)
    pkg_headers = ['Paket', 'Tagline', 'İçerik', 'Bütçe']
    for col, h in enumerate(pkg_headers, 1):
        c = ws2.cell(row=5, column=col, value=h)
        c.font = Font(bold=True, color="181818", size=11)
        c.fill = PatternFill("solid", fgColor="d4a843")
        c.alignment = Alignment(horizontal="left", vertical="center")
    pkg_rows = [
        ('BRONZ', '"Görünürlük"', 'Altbant + Ürün Yerleştirme', '₺1.250.000 +KDV'),
        ('ALTIN', '"Sahiplen"', 'Sundu + Altbant + Ürün Yerleştirme', '₺2.320.000 +KDV'),
        ('ÖZEL', '"Konuş & Sahiplen"', 'Entegrasyon + Sundu', '₺3.350.000 +KDV'),
        ('ELMAS', '"Tam Kapsama"',
         'Entegrasyon + Sundu + Altbant + Ürün Yerleştirme', '₺3.970.000 +KDV'),
    ]
    for i, row in enumerate(pkg_rows, start=6):
        ws2.cell(row=i, column=1, value=row[0]).font = Font(bold=True, size=12)
        ws2.cell(row=i, column=2, value=row[1]).font = Font(italic=True)
        ws2.cell(row=i, column=3, value=row[2]).alignment = Alignment(wrap_text=True, vertical="top")
        c_new = ws2.cell(row=i, column=4, value=row[3])
        c_new.font = Font(bold=True, color="d4a843", size=12)
        c_new.alignment = Alignment(horizontal="right")
        ws2.row_dimensions[i].height = 30
    _excel_columns(ws2, [25, 22, 55, 22])
    _excel_borders(ws2, 5, 9, 1, 4)

    wb.save(str(out_path))


def _excel_header(ws, title_text, n_cols=5):
    """Her iki logo SOLDA yan yana: Kırmızı-Transparan (D20) + Mavi-Crossover-Sports."""
    # Kırmızı-Transparan: 1200x960 ≈ 5:4
    if LOGO_KIRMIZI.exists():
        try:
            img = XLImage(str(LOGO_KIRMIZI))
            img.width = 95
            img.height = 76          # 5:4 aspect
            ws.add_image(img, "A1")
        except Exception:
            pass
    # Mavi-Crossover: 1080x1080 = 1:1 (kare)
    if LOGO_MAVI.exists():
        try:
            img = XLImage(str(LOGO_MAVI))
            img.width = 75
            img.height = 75          # 1:1 aspect — sıkışmayı önler
            ws.add_image(img, "B1")  # Kırmızı'nın hemen sağında
        except Exception:
            pass
    for r in range(1, 5):
        ws.row_dimensions[r].height = 22
    # İlk iki kolonu logolar için yeterince geniş tut
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 13
    ws.cell(row=4, column=1, value=title_text).font = Font(bold=True, color="d4a843", size=13)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)


def _excel_columns(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _excel_borders(ws, row_start, row_end, col_start, col_end):
    thin = Side(border_style="thin", color="cccccc")
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ─── CLI ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    pdf_path = ASSETS / "kupa-sunum.pdf"
    xlsx_path = ASSETS / "kupa-paketler.xlsx"
    print("Building PDF...")
    build_pdf(pdf_path)
    print(f"  OK  {pdf_path.relative_to(ROOT)}  ({pdf_path.stat().st_size:,} bytes)")
    print("Building Excel...")
    build_excel(xlsx_path)
    print(f"  OK  {xlsx_path.relative_to(ROOT)}  ({xlsx_path.stat().st_size:,} bytes)")
